import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import os

STANDARD_FILE = r"D:\dev\auto_organon\standard.xlsx"
BASE_PATH = r"D:\dev\auto_organon\cases\auto-organon_output\base"
SENSITIVITY_PATH = r"D:\dev\auto_organon\cases\auto-organon_output\sensitivity"

class Sheet:
    def __init__(self, sheet_name, start_row, start_col, start_col_sensitivity, sort_col, ascending):
        self.sheet_name = sheet_name
        self.start_row = start_row
        self.start_col = start_col
        self.start_col_sensitivity = start_col_sensitivity
        self.sort_col = sort_col
        self.ascending = ascending

sheets = [
    Sheet("PWF03", 3, 1, 10, " Volt(pu)", True),
    Sheet("PWF05", 3, 1, 10, " VMin(pu)", False),
    Sheet("PWF16", 3, 2, 21, " % L1", False),
    Sheet("CTG01", 3, 2, 15, " Viol (pu)", False),
    Sheet("CTG02", 3, 2, 16, " Viol (pu)", False),
    Sheet("CTG03", 3, 2, 19, " Viol (%)", False),
]

for folder_name in os.listdir(BASE_PATH):
    folder_path = os.path.join(BASE_PATH, folder_name)
    if os.path.isdir(folder_path):
        excel_file = os.path.join(folder_path, folder_name + ".xlsx")
        shutil.copy(STANDARD_FILE, excel_file)

        workbook = load_workbook(excel_file)

        for sheet in sheets:
            sheet_name = sheet.sheet_name
            start_row = sheet.start_row
            start_col = sheet.start_col
            sort_col = sheet.sort_col
            ascending = sheet.ascending

            csv_file = os.path.join(folder_path, sheet_name + ".csv")
            df = pd.read_csv(csv_file, encoding="latin-1", sep=";", decimal=".", skiprows=1)
            df.sort_values(by=sort_col, ascending=ascending, inplace=True)

            sheet = workbook[sheet_name]
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
                for c_idx, value in enumerate(row, start=start_col):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

        workbook.save(excel_file)

for folder_name in os.listdir(SENSITIVITY_PATH):
    folder_path = os.path.join(SENSITIVITY_PATH, folder_name)
    if os.path.isdir(folder_path):
        BASE_FILE = os.path.join(BASE_PATH, folder_name, folder_name + ".xlsx")
        excel_file = os.path.join(folder_path, folder_name + ".xlsx")
        shutil.copy(BASE_FILE, excel_file)

        workbook = load_workbook(excel_file)

        for sheet in sheets:
            sheet_name = sheet.sheet_name
            start_row = sheet.start_row
            start_col_sensitivity = sheet.start_col_sensitivity
            sort_col = sheet.sort_col
            ascending = sheet.ascending

            csv_file = os.path.join(folder_path, sheet_name + ".csv")
            df = pd.read_csv(csv_file, encoding="latin-1", sep=";", decimal=".", skiprows=1)
            df.sort_values(by=sort_col, ascending=ascending, inplace=True)

            sheet = workbook[sheet_name]
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
                for c_idx, value in enumerate(row, start=start_col_sensitivity):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

        workbook.save(excel_file)