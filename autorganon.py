import csv
import os
import shutil
import subprocess
import math
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

CONFIG_FILE = "configs.json"
DEFAULT_VALUES = {
    "case_path": r"D:\dev\auto_organon\cases",
    "line_file": "2028Lines.csv",
    "trf_file": "2028Transfo.csv",
    "bus": "7190",
    "levels": "3",
    "areas": "773, 772, 771, 761",
    "cases_file": "cases.csv"
}

class Sheet:
    def __init__(self, sheet_name, start_row, start_col, start_col_sensitivity, sort_col, ascending):
        self.sheet_name = sheet_name
        self.start_row = start_row
        self.start_col = start_col
        self.start_col_sensitivity = start_col_sensitivity
        self.sort_col = sort_col
        self.ascending = ascending

sheets = [
    Sheet("PWF03", 3, 1, 10, " Volt(pu)", True),
    Sheet("PWF05", 3, 1, 10, " VMin(pu)", False),
    Sheet("PWF16", 3, 2, 23, " % L1", False),
    Sheet("CTG01", 3, 2, 15, " Viol (pu)", False),
    Sheet("CTG02", 3, 2, 16, " Viol (pu)", False),
    Sheet("CTG03", 3, 2, 19, " Viol (%)", False),
]


class Line:
    def __init__(self, row):
        self.from_bus, self.to_bus, self.circ, self.from_name, self.to_name = extract_circuit_data(row)

    def __str__(self):
        return f"<{self.__class__} {self.from_bus} {self.to_bus} {self.circ}>"

    def __repr__(self):
        return self.__str__()

    def __eq__(self, other):
        return (self.from_bus == other.from_bus and self.to_bus == other.to_bus) or \
            (self.from_bus == other.to_bus and self.to_bus == other.from_bus)

    def __hash__(self):
        return self.from_bus + 10000 * self.to_bus


class Transformer(Line):
    pass


def load_values(config_file="configs.json"):
    with open(config_file, "r") as file:
        return json.load(file)


def save_values(updated_values):
    values = load_values()
    values.update(updated_values)
    with open(CONFIG_FILE, "w") as file:
        json.dump(values, file, indent=4)


def extract_circuit_data(row: list) -> tuple:
    from_bus = int(row[0])
    from_name = row[1].strip()
    to_bus, circ = tuple(map(lambda x: int(x.strip()), row[2].split("#")))
    to_name = row[3]

    return from_bus, to_bus, circ, from_name, to_name


def get_circuits_connected_to(circuits: list, bus_number: int) -> list:
    captured = []
    for circuit in circuits:
        if circuit.from_bus == bus_number or circuit.to_bus == bus_number:
            if not circuit in captured:
                captured.append(circuit)
    return captured


def get_unique_bus_numbers(circuits: list) -> list:
    buses = set()
    for circuit in circuits:
        buses.add(circuit.from_bus)
        buses.add(circuit.to_bus)
    return list(buses)


def extract_voltage_value(name: str) -> int:
    for length in range(3, 0, -1):
        try:
            return int(name[-length:])
        except ValueError:
            continue
    return None


def get_neighboring_circuits(circuits: list, central_bus: int, levels: int) -> list:
    captured = set()
    buses_to_visit = [central_bus, ]
    already_visited = []
    while levels >= 1:
        new_circuits = set()
        for bus in buses_to_visit:
            if bus not in already_visited:
                connected_circuits = get_circuits_connected_to(circuits, bus)
                for circuit in connected_circuits:
                    from_voltage = extract_voltage_value(circuit.from_name)
                    to_voltage = extract_voltage_value(circuit.to_name)

                    if (from_voltage is None or from_voltage >= 230 or from_voltage == 0) and \
                            (to_voltage is None or to_voltage >= 230 or to_voltage == 0):
                        new_circuits.add(circuit)

                already_visited.append(bus)

        buses_to_visit = get_unique_bus_numbers(new_circuits)
        captured.update(new_circuits)
        levels = levels - 1
    return list(captured)


def gen_ctg(CASE_PATH, line_file_name, trf_file_name, input_bus, input_levels):
    lines = []
    transformers = []

    with open(os.path.join(CASE_PATH, line_file_name), "r") as lines_file:
        reader = csv.reader(lines_file, delimiter=";")
        next(reader)
        next(reader)
        for row in reader:
            if len(row) > 1:
                lines.append(Line(row))

    with open(os.path.join(CASE_PATH, trf_file_name), "r") as trf_file:
        reader = csv.reader(trf_file, delimiter=";")
        next(reader)
        next(reader)
        for row in reader:
            if len(row) > 1:
                transformers.append(Transformer(row))

    circuits = []
    circuits.extend(lines)
    circuits.extend(transformers)

    first_circuits = get_circuits_connected_to(circuits, input_bus)
    buses = get_unique_bus_numbers(first_circuits)

    circs = get_neighboring_circuits(circuits, input_bus, input_levels)

    organon_ctg = """  '{name:60s}' 
	BRANCH      {from_b:5d}        {to_b:5d}       {circ:02d}
			END /
	"""

    with open(os.path.join(CASE_PATH, "contigencies.ctg"), "w") as file:
        for circ in circs:
            file.write(
                organon_ctg.format(name=f"{circ.from_name}({circ.from_bus}) {circ.to_name}({circ.to_bus}) #{circ.circ}",
                                   from_b=circ.from_bus, to_b=circ.to_bus, circ=circ.circ))


def gen_spt(CASE_PATH, cases_file_name="cases.csv", RUN_CONTINGENCY=True):
    CASES_FILE = os.path.join(CASE_PATH, cases_file_name)
    OUTPUT_SCRIPT = os.path.join(CASE_PATH, "script.spt")
    DEF_FILE = os.path.join(CASE_PATH, "definitions.def")
    CTG_FILE = os.path.join(CASE_PATH, "contigencies.ctg")
    OUTPUT_PATH = os.path.join(CASE_PATH, "output")

    input_file_paths = []
    with open(CASES_FILE, "r") as case_file:
        for row in csv.reader(case_file):
            file_path = os.path.join(CASE_PATH, row[0])

            if os.path.exists(file_path):
                input_file_paths.append(file_path)
            else:
                print(f"File does not exist: {file_path}")

    spt_lines = []

    for ifile, file_path in enumerate(input_file_paths):
        base_name, ext = os.path.splitext(os.path.basename(file_path))
        output_folder_path = os.path.join(OUTPUT_PATH, base_name)
        os.makedirs(output_folder_path, exist_ok=True)
        output_file_path = os.path.join(output_folder_path, base_name + ext)
        shutil.copyfile(file_path, output_file_path)

        abs_file_path = os.path.abspath(output_file_path)
        spt_lines.append(f"OPEN \"{abs_file_path}\"")
        spt_lines.append(f"OPEN \"{DEF_FILE}\"")
        spt_lines.append(f"OPEN \"{CTG_FILE}\"")
        spt_lines.append("")
        spt_lines.append("NEWTON")
        spt_lines.append("CSV PWF03")
        spt_lines.append("CSV PWF05")
        spt_lines.append("CSV PWF16")
        spt_lines.append("")
        if RUN_CONTINGENCY:
            spt_lines.append("PFCTG")
            spt_lines.append("CSV CTG01")
            spt_lines.append("CSV CTG02")
            spt_lines.append("CSV CTG03")
        spt_lines.append("")
        spt_lines.append("SAVE PFD")

        if ifile < len(input_file_paths) - 1:
            spt_lines.append("")
            spt_lines.append("! ---")
            spt_lines.append("")

    with open(OUTPUT_SCRIPT, "w") as spt_file:
        spt_file.write("\n".join(spt_lines))


def gen_def(areas, CASE_PATH):
    monitor_template = " AREA      {number:5d}     230.0  999.0 ONE  /\n"

    with open(os.path.join(CASE_PATH, "definitions.def"), "w") as file:
        file.write(" MONITOR\n")
        for area in areas:
            line = monitor_template.format(number=area)
            file.write(line)

        file.write(" END /\n")


def run_organon(script_path):
    executable_path = r"C:\Program Files\HPPA\Organon\bin\OrgProc.exe"
    arguments = ['1', script_path, '1']
    command = [executable_path] + arguments

    try:
        process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

        while True:
            output = process.stdout.readline()
            if output == '' and process.poll() is not None:
                break
            if output:
                print(output.strip())

        stderr = process.stderr.read()
        if stderr:
            print("STDERR:", stderr.strip())

        process.stdout.close()
        process.stderr.close()

        if process.wait() != 0:
            print("Execution failed with return code:", process.returncode)

    except Exception as e:
        print("An error occurred while running the Organon process:", e)


def format_number(number):
    if number >= 10000:
        formatted = f"{number:.0f}"
    elif number >= 1000:
        formatted = f"{number:.1f}"
    elif number >= 100:
        formatted = f"{number:.2f}"
    elif number >= 10:
        formatted = f"{number:.3f}"
    else:
        formatted = f"{number:.4f}"

    return formatted[:5]


def add_load(pl, bus, pf, folder_path):
    ql = pl * math.tan(math.acos(pf))
    for subdir, dirs, files in os.walk(folder_path):
        for file_name in files:
            if file_name.endswith('.PWF'):
                file_path = os.path.join(subdir, file_name)

                with open(file_path, 'r', encoding='latin-1') as file:
                    content = file.readlines()

                section_started = False
                for i, line in enumerate(content):
                    if line.strip() == 'DBAR':
                        section_started = True
                        continue
                    if line.strip()[0] == '(':
                        continue
                    if section_started:
                        if line.strip() == '99999':
                            break
                        line_parts = line.split()
                        if int(line_parts[0]) == bus:
                            existing_Pl = float(line[58:63].strip()) if line[58:63].strip() else 0.0
                            existing_Ql = float(line[63:68].strip()) if line[63:68].strip() else 0.0

                            updated_Pl = existing_Pl + pl
                            updated_Ql = existing_Ql + ql

                            new_line = (line[
                                        :58] + f"{format_number(updated_Pl):>5}" + f"{format_number(updated_Ql):>5}" + line[
                                                                                                                       68:])
                            content[i] = new_line

                with open(file_path, 'w', encoding='latin-1') as file:
                    file.writelines(content)

                print(f"The file {file_path} has been successfully updated.")


def compare_cases(base_path, sensitivity_path):
    BASE_PATH = os.path.join(base_path, "output")
    SENSITIVITY_PATH = os.path.join(sensitivity_path, "output")
    STANDARD_FILE = os.path.join(base_path, "standard.xlsx")
    for folder_name in os.listdir(BASE_PATH):
        folder_path = os.path.join(BASE_PATH, folder_name)
        if os.path.isdir(folder_path):
            excel_file = os.path.join(folder_path, folder_name + ".xlsx")
            shutil.copy(STANDARD_FILE, excel_file)

            workbook = load_workbook(excel_file)

            for sheet in sheets:
                sheet_name = sheet.sheet_name
                start_row = sheet.start_row
                start_col = sheet.start_col
                sort_col = sheet.sort_col
                ascending = sheet.ascending

                csv_file = os.path.join(folder_path, sheet_name + ".csv")
                df = pd.read_csv(csv_file, encoding="latin-1", sep=";", decimal=".", skiprows=1)
                df.sort_values(by=sort_col, ascending=ascending, inplace=True)

                sheet = workbook[sheet_name]
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
                    for c_idx, value in enumerate(row, start=start_col):
                        sheet.cell(row=r_idx, column=c_idx, value=value)

            workbook.save(excel_file)

    for folder_name in os.listdir(SENSITIVITY_PATH):
        folder_path = os.path.join(SENSITIVITY_PATH, folder_name)
        if os.path.isdir(folder_path):
            BASE_FILE = os.path.join(BASE_PATH, folder_name, folder_name + ".xlsx")
            excel_file = os.path.join(folder_path, folder_name + ".xlsx")
            shutil.copy(BASE_FILE, excel_file)

            workbook = load_workbook(excel_file)

            for sheet in sheets:
                sheet_name = sheet.sheet_name
                start_row = sheet.start_row
                start_col_sensitivity = sheet.start_col_sensitivity
                sort_col = sheet.sort_col
                ascending = sheet.ascending

                csv_file = os.path.join(folder_path, sheet_name + ".csv")
                df = pd.read_csv(csv_file, encoding="latin-1", sep=";", decimal=".", skiprows=1)
                df.sort_values(by=sort_col, ascending=ascending, inplace=True)

                sheet = workbook[sheet_name]
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
                    for c_idx, value in enumerate(row, start=start_col_sensitivity):
                        sheet.cell(row=r_idx, column=c_idx, value=value)

            workbook.save(excel_file)


def save_values(updated_values):
    values = load_values()
    values.update(updated_values)
    with open(CONFIG_FILE, "w") as file:
        json.dump(values, file, indent=4)


def run_gen_ctg(config):
    try:
        bus = int(config.get("bus", 0))
        levels = int(config.get("levels", 0))
        gen_ctg(config.get("case_path"), config.get("line_file"), config.get("trf_file"), bus, levels)
        print(".ctg generated successfully!")
    except ValueError as e:
        print(f"Error in .ctg generation: {e}")


def run_gen_def(config):
    try:
        areas = [int(area.strip()) for area in config.get("areas", "").split(',')]
        gen_def(areas, config.get("case_path"))
        print(".def generated successfully!")
    except ValueError as e:
        print(f"Error in .def generation: {e}")


def run_gen_spt(config):
    try:
        gen_spt(config.get("case_path"), config.get("cases_file"))
        print(".spt generated successfully!")
    except Exception as e:
        print(f"Error in .spt generation: {e}")